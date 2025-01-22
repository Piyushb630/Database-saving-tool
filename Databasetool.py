import customtkinter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import messagebox, ttk
#from win32com.client import Dispatch
import os
import pandas as pd
 

# Set appearance mode and default color theme for customtkinter
customtkinter.set_appearance_mode("system")
# customtkinter.set_default_color_theme("blue")

excel_path = r'Location of excel'
 
Organization_email = "xyz@gmail.com"



def refresh(excel_path):
  process_excel(excel_path)
  messagebox.showinfo("Info",f"Data has been refreshed successfully")


def process_excel(excel_path):
    try:
        workbook = load_workbook(excel_path, data_only=True)
        if "Database" in workbook.sheetnames:
            sheet = workbook["Database"]
        else:
            messagebox.showerror("Error", "Sheet named 'Database' does not exist in the Excel file.")
            return
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open Excel file: {str(e)}")
        return

    try:
        # Extract data and highlight info
        data = []
        highlights = []
        for row in sheet.iter_rows(min_row=2):  # Skip header row
            data_row = [cell.value for cell in row]
            highlight_row = [
                cell.fill.start_color.rgb if cell.fill and cell.fill.start_color.rgb != '00000000' else None 
                for cell in row
            ]
            data.append(data_row)
            highlights.append(highlight_row)
        
        # Create DataFrame for display
        df = pd.DataFrame(data, columns=[cell.value for cell in sheet[1]])
        
        # Pass both data and highlights to show_data
        show_data(tab2_tree, df, highlights)
        
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while processing the Excel file: {str(e)}")

def show_data(tree, df, highlights):
    tree.delete(*tree.get_children())
    
    columns = list(df.columns)
    tree["columns"] = columns
    
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=100)
    
    for index, (row, highlight_row) in enumerate(zip(df.iterrows(), highlights)):
        tree_row = tree.insert("", "end", values=list(row[1]), tags=(f"row{index}",))
        
        # Apply colors to cells based on highlight info
        for col_index, color in enumerate(highlight_row):
            cell_tag = f"cell{index}_{col_index}"
            if color:
                # Apply tag with color
                tree.tag_configure(cell_tag, background=f"#{color[2:]}")
            else:
                # Apply default color
                tree.tag_configure(cell_tag, background="#FFFFFF")
                
            # Assign tag to cell
            tree.item(tree_row, tags=(f"row{index}", cell_tag))
    
    # Configure default style for treeview (optional)
    tree.tag_configure("default", background="#FFFFFF")
   


def get_outlook_signature():
    global signature
    # Create a temporary email to extract the default signature
    outlook = Dispatch('Outlook.Application')
    mail = outlook.CreateItem(0)  # Create a new email item
    mail.GetInspector
    signature = mail.HTMLBody
    return signature

def generate_salesapprovalemail():
    client_name = client_input.get()
    pb_name = pb_input.get()
    if not client_name or not pb_name:
        messagebox.showwarning("Input Error", "Please input client's name and PB name before creating email.")
        return

    outlook = Dispatch('Outlook.Application')
    email = outlook.CreateItem(0)
    email.Subject = f"New FX give-up Prospect - {client_name}"
    email.SentOnBehalfOfName = Organization_email
    email.CC = Organization_email
    email.To = ""
    email.Body = f"Body of the email."
    email.Display(False)


def generate_onboardingemail():
    client_name = client_input.get()
    pb_name = pb_input.get()
    if not client_name or not pb_name:
        messagebox.showwarning("Input Error", "Please input client's name and PB name before creating email.")
        return

    outlook = Dispatch('Outlook.Application')
    email = outlook.CreateItem(0)
    email.Subject = f"New FX give-up - {client_name}"
    email.SentOnBehalfOfName = Organization_email
    email.CC = Organization_email
    email.To = ""
    email.Body = f"Body of the email"
    email.Display(False)


def USP():
    client_name = client_input.get()
    pb_name = pb_input.get()
    if not client_name or not pb_name:
        messagebox.showwarning("Input Error", "Please input client's name and PB name before creating email.")
        return

    outlook = Dispatch('Outlook.Application')
    email = outlook.CreateItem(0)
    email.Subject = f"USP rider status query for - {client_name}"
    email.SentOnBehalfOfName = Organization_email
    email.CC = Organization_email
    email.To = ""  # Replace with the actual recipient's email address
    email.Body = f"Hello,\n\nCould you please advise if the client {client_name} is a USP?\n\n If the Agent is not a US Person, please provide an email where the Agent or PB confirm whether they are a US Person."
    email.Display(False)


def generate_limitsemail():
    client_name = client_input.get()
    pb_name = pb_input.get()
    gcn = GCN_input.get()
    if not client_name or not pb_name or not gcn:
        messagebox.showwarning("Input Error", "Please input client's name,PB name and GCN before creating email.")
        return

   

    outlook = Dispatch('Outlook.Application')
    email = outlook.CreateItem(0)
    email.Subject = f"Update limit for - {pb_name} {client_name}/{gcn}"
    email.SentOnBehalfOfName = Organization_email
    email.CC = Organization_email
    email.To = ""
    email.Body = f"Hello team,\n\nPlease update the limits for {pb_name}/{client_name} {gcn} according to the attached EDN.\n\n Please confirm once processed."
    email.Display(False)


def generate_agreementsemail():
    global value_d, value_e

    pb_name = pb_input.get()
    client_name = client_input.get()
    gcn = GCN_input.get()
    if not client_name or not pb_name or not gcn:
        messagebox.showwarning("Input Error", "Please input client's name,PB name and GCN before creating email.")
        return

 

    try:
        workbook = load_workbook(excel_path)
        sheet = workbook["PB_GCN"]
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open Excel file: {str(e)}")
        return  # Exit if the Excel file cannot be opened

    found_row = None
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value == pb_name:
            found_row = row
            value_d = sheet.cell(row, 3).value  # Column C (3rd column)
   
            break  # Exit the loop once the row is found

    if found_row is None:
        messagebox.showinfo("Not Found", f"PB Name '{pb_name}' not found in the Excel sheet.")
        return  # Exit if pb_name is not found

    # Create and send the email
    outlook = Dispatch('Outlook.Application')
    email = outlook.CreateItem(0)
    email.Subject = f"FX Give up {pb_name}/{client_name}"
    email.SentOnBehalfOfName = Organization_email
    email.CC = Organization_email
    email.To = ""
    email.Body = (
        f"Hello team,\n\n"
        f"Please include GCN: {pb_name}/{client_name}: {gcn} Body of the mail{pb_name}: {value_d}.\n\n"
        f"Please confirm once processed.\n"
    )
    email.Display(False)  # Display the email before sending

def store_in_excel():
    try:
        workbook = load_workbook(excel_path, data_only=True)
        if "Database" in workbook.sheetnames:
            sheet = workbook["Database"]
        else:
            messagebox.showerror("Error", "Sheet named 'Database' does not exist in the Excel file.")
            return
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open Excel file: {str(e)}")
        return

    # Get values from UI elements
    client_name = client_input.get()
    pb_name = pb_input.get()
    approval_from_sales_CX = checkbox_approval_from_sales_CX.get()
    usp_rider_status = checkbox_UPSRIDER_status.get()
    email_onboarding = checkbox_inintitate_onboarding.get()
    setup_murex_calypso = checkbox_GCN_setup_calypso_murex.get()
    limits_in_place = checkbox_Limits.get()
    agreement_in_place = checkbox_Agreements.get()
    GCN = GCN_input.get()
    Shortname = Shortname_input.get()
    ERF_setup = checkbox_ERF_setup.get()
    EDN_ID = EDN_input.get()

    if not client_name or not pb_name or not EDN_ID:
        messagebox.showwarning("Input Error", "Please input client's name and PB name before storing data.")
        return


    # Find the next empty row or update existing row based on conditions
    found_row = None
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 1).value == client_name and sheet.cell(row, 2).value == pb_name:
            found_row = row
            overwrite = messagebox.askyesno("Duplicate Entry", f"Duplicate found for Client Name '{client_name}' and PB Name '{pb_name}'. Do you want to overwrite?")
            if not overwrite :
                return  # If the user selects "No", exit the function
            break

    if found_row:
        # Update existing row
        sheet.cell(found_row, 3).value = 'Yes' if approval_from_sales_CX else 'No'
        sheet.cell(found_row, 4).value = 'Yes' if usp_rider_status else 'No'
        sheet.cell(found_row, 5).value = 'Yes' if email_onboarding else 'No'
        sheet.cell(found_row, 6).value = 'Yes' if setup_murex_calypso else 'No'
        sheet.cell(found_row, 7).value = 'Yes' if limits_in_place else 'No'
        sheet.cell(found_row, 8).value = 'Yes' if agreement_in_place else 'No'
        sheet.cell(found_row, 9).value = 'Yes' if ERF_setup else 'No'
        sheet.cell(found_row, 10).value = GCN
        sheet.cell(found_row, 11).value = Shortname
        sheet.cell(found_row, 12).value = EDN_ID
    else:
        # Insert new row
        next_row = sheet.max_row + 1
        sheet.cell(next_row, 1).value = client_name
        sheet.cell(next_row, 2).value = pb_name
        sheet.cell(next_row, 3).value = 'Yes' if approval_from_sales_CX else 'No'
        sheet.cell(next_row, 4).value = 'Yes' if usp_rider_status else 'No'
        sheet.cell(next_row, 5).value = 'Yes' if email_onboarding else 'No'
        sheet.cell(next_row, 6).value = 'Yes' if setup_murex_calypso else 'No'
        sheet.cell(next_row, 7).value = 'Yes' if limits_in_place else 'No'
        sheet.cell(next_row, 8).value = 'Yes' if agreement_in_place else 'No'
        sheet.cell(next_row, 9).value = 'Yes' if ERF_setup else 'No'
        sheet.cell(next_row, 10).value = GCN
        sheet.cell(next_row, 11).value = Shortname
        sheet.cell(next_row, 12).value = EDN_ID

    # Save workbook
    try:
        workbook.save(excel_path)
        messagebox.showinfo("Success", "Data stored successfully in Excel.")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save data to Excel: {str(e)}")


def terminate():

    try:
        workbook = load_workbook(excel_path)
        if "Database" in workbook.sheetnames:
            sheet = workbook["Database"]
        else:
            messagebox.showerror("Error", "Sheet named 'Database' does not exist in the Excel file.")
            return
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open Excel file: {str(e)}")
        return

    client_name = client_input.get()
    pb_name = pb_input.get()
    approval_from_sales_CX = checkbox_approval_from_sales_CX.get()
    usp_rider_status = checkbox_UPSRIDER_status.get()
    email_onboarding = checkbox_inintitate_onboarding.get()
    setup_murex_calypso = checkbox_GCN_setup_calypso_murex.get()
    limits_in_place = checkbox_Limits.get()
    agreement_in_place = checkbox_Agreements.get()
    GCN = GCN_input.get()
    Shortname = Shortname_input.get()
    ERF_setup = checkbox_ERF_setup.get()
    EDN_ID = EDN_input.get()
    found_row = None

    if not client_name or not pb_name or not EDN_ID:
        messagebox.showwarning("Input Error", "Please input client's name and PB name before storing data.")
        return


    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 1).value == client_name and sheet.cell(row, 2).value == pb_name:
            found_row = row
            overwrite = messagebox.askyesno("Entry", f"Entry found for Client Name '{client_name}' and PB Name '{pb_name}'. Do you want to Terminate?")
            if not overwrite :
                return  # If the user selects "No", exit the function
            break

    if found_row:
        # Highlight the entire row in red
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=found_row, column=col).fill = red_fill

        # Optionally save the workbook after highlighting the row
        workbook.save(excel_path)
        messagebox.showinfo("Row Highlighted", "The duplicate row has been highlighted in red.")

    else:
        # If no duplicate is found, you can add code to handle this case, if needed
        messagebox.showinfo("No Duplicate Found", "No duplicate entry was found.")

    pb_name = pb_input.get()
    client_name = client_input.get()
    gcn = GCN_input.get()

    try:
        workbook = load_workbook(excel_path)
        sheet = workbook["PB_GCN"]
    except Exception as e:
        messagebox.showerror("Error", f"Failed to open Excel file: {str(e)}")
        return  # Exit if the Excel file cannot be opened

    found_row = None
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value == pb_name:
            found_row = row
            value_d = sheet.cell(row, 3).value  # Column C (3rd column)
   
            break  # Exit the loop once the row is found

    if found_row is None:
        messagebox.showinfo("Not Found", f"PB Name '{pb_name}' not found in the Excel sheet.")
        return  # Exit if pb_name is not found

    # Create and send the email
    outlook = Dispatch('Outlook.Application')
    email = outlook.CreateItem(0)
    email.Subject = f"FX Give up {pb_name}/{client_name}"
    email.SentOnBehalfOfName = Organization_email
    email.CC = Organization_email
    email.To = "marketsagreementsmanagement@seb.se; FICoverageBanksCAN@seb.se "
    email.Body = (
        f"Hello team,\n\n"
        f"Please remove the limits for {pb_name}/{client_name} {gcn} according to the attached EDN.\n\n"
        f"Please remove GCN: {pb_name}/{client_name}: {gcn} as an IC to the ISDA and CSA for GCN {pb_name}: {value_d}.\n\n"
        f"Please confirm once processed.\n"
    )
    email.Display(False)  # Display the email before sending

def on_treeview_select(event):
    selection = tab2_tree.selection()
    if selection:
        item = tab2_tree.item(selection[0])
        client_name = item['values'][0]
        pb_name = item['values'][1]
        # Retrieve other values from the Excel file
        workbook = load_workbook(excel_path)
        sheet = workbook["Database"]
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 1).value == client_name:
                GCN = sheet.cell(row, 10).value
                Shortname = sheet.cell(row, 11).value
                EDN_ID = sheet.cell(row, 12).value
                break
        # Fill in the fields in tab1
        client_input.delete(0, 'end')
        client_input.insert(0, client_name)
        pb_input.set(pb_name)
        GCN_input.delete(0, 'end')
        GCN_input.insert(0, str(GCN))  # Convert GCN to string
        Shortname_input.delete(0, 'end')
        Shortname_input.insert(0, Shortname)
        EDN_input.delete(0, 'end')
        EDN_input.insert(0, str(EDN_ID))  # Convert EDN_ID to string
        
def open_excel():
    os.startfile(excel_path)

window = customtkinter.CTk()
window.geometry("900x900")
window.title("CRM Process")

tab_control = customtkinter.CTkTabview(window)
tab1 = tab_control.add("CRM DATABASE TOOL")
tab2 = tab_control.add("DATABASE")
tab_control.pack(expand=1, fill="both")

tab1_frame = customtkinter.CTkFrame(master=tab1)
tab1_frame.pack(pady=20, padx=60, fill="both", expand=True)

label_frame = customtkinter.CTkFrame(master=tab1_frame)
label_frame.pack(pady=12, padx=10)


heading = customtkinter.CTkLabel(master=label_frame, text="CRM DATABASE TOOL", width=400)
heading.pack(pady=12, padx=10)

client_input = customtkinter.CTkEntry(master=tab1_frame, placeholder_text="Input client's name", width=200)
client_input.pack(pady=12, padx=10)

pb_names = ["Bank1","Bank2","Bank3"]
pb_input = customtkinter.CTkComboBox(master=tab1_frame, values=pb_names, width=200)
pb_input.pack(pady=12, padx=10)

GCN_input = customtkinter.CTkEntry(master=tab1_frame, placeholder_text="Input GCN", width=200)
GCN_input.pack(pady=12, padx=10)

Shortname_input = customtkinter.CTkEntry(master=tab1_frame, placeholder_text="Input Shortname", width=200)
Shortname_input.pack(pady=12, padx=10)

EDN_input = customtkinter.CTkEntry(master=tab1_frame, placeholder_text="EDN ID/PAPER")
EDN_input.pack(pady=12, padx=10)

Button1_frame = customtkinter.CTkFrame(master=tab1_frame)
Button1_frame.pack(pady=12, padx=10)

generate_email_button = customtkinter.CTkButton(master=Button1_frame, text="Generate approval email from sales and CX", width=300, command=generate_salesapprovalemail)
generate_email_button.grid(row=0, column=0, sticky='w', padx=20, pady=10)

generate_USPEMAILtoPB = customtkinter.CTkButton(master=Button1_frame, text="Generate USP rider status email to PB", width=300, command=USP)
generate_USPEMAILtoPB.grid(row=0, column=1, sticky='w', padx=20, pady=10)

onboarding_email_button = customtkinter.CTkButton(master=Button1_frame, text="Generate onboarding email", width=300, command=generate_onboardingemail)
onboarding_email_button.grid(row=1, column=0, sticky='w', padx=20, pady=10)

generate_emailforlimits = customtkinter.CTkButton(master=Button1_frame, text="Email to CCM banks for limits", width=300, command=generate_limitsemail)
generate_emailforlimits.grid(row=1, column=1, sticky='w', padx=20, pady=10)

generate_emailforagreements = customtkinter.CTkButton(master=Button1_frame, text="Email to add agreement", width=300, command=generate_agreementsemail)
generate_emailforagreements.grid(row=2, column=0, sticky='w', padx=20, pady=10)

checkbox_frame = customtkinter.CTkFrame(master=tab1_frame)
checkbox_frame.pack(pady=12, padx=10)

checkbox_approval_from_sales_CX = customtkinter.CTkCheckBox(master=checkbox_frame, text="Approval from sales and CX")
checkbox_approval_from_sales_CX.grid(row=0, column=0, sticky='w', padx=20, pady=10)

checkbox_UPSRIDER_status = customtkinter.CTkCheckBox(master=checkbox_frame, text="USP rider status received")
checkbox_UPSRIDER_status.grid(row=0, column=1, sticky='w', padx=20, pady=10)

checkbox_inintitate_onboarding = customtkinter.CTkCheckBox(master=checkbox_frame, text="Email sent for onboarding")
checkbox_inintitate_onboarding.grid(row=1, column=0, sticky='w', padx=20, pady=10)

checkbox_GCN_setup_calypso_murex = customtkinter.CTkCheckBox(master=checkbox_frame, text="Setup is there in Calypso/Murex")
checkbox_GCN_setup_calypso_murex.grid(row=1, column=1, sticky='w', padx=20, pady=10)

checkbox_ERF_setup = customtkinter.CTkCheckBox(master=checkbox_frame, text="ERF setup")
checkbox_ERF_setup.grid(row=2, column=0, sticky='w', padx=20, pady=10)

checkbox_Limits = customtkinter.CTkCheckBox(master=checkbox_frame, text="Limits updated")
checkbox_Limits.grid(row=2, column=1, sticky='w', padx=20, pady=10)

checkbox_Agreements = customtkinter.CTkCheckBox(master=checkbox_frame, text="Agreements in place")
checkbox_Agreements.grid(row=3, column=0, sticky='w', padx=20, pady=10)

button_frame = customtkinter.CTkFrame(master=tab1_frame)
button_frame.pack(pady=12, padx=10)

store_button = customtkinter.CTkButton(master=button_frame, text="NEW/UPDATE DATABASE", command=store_in_excel)
store_button.grid(row=0, column=0, sticky='w', padx=20, pady=10)

Terminatebutton = customtkinter.CTkButton(master=button_frame, text ="TERMINATE IN DATABASE", command=terminate)
Terminatebutton.grid(row=0, column=1, sticky='w', padx=20, pady=10)

view_database_button = customtkinter.CTkButton(master=button_frame, text="VIEW DATABASE", command=open_excel)
view_database_button.grid(row=0, column=2, sticky='w', padx=20, pady=10)


tab2_frame = customtkinter.CTkFrame(master=tab2)
tab2_frame.pack(pady=20, padx=60, fill="both", expand=True)

tab2_tree = ttk.Treeview(tab2_frame, show='headings', selectmode='extended')
tab2_tree.pack(expand=1, fill='both')

tab2_tree.bind('<<TreeviewSelect>>', on_treeview_select)

button_refresh = customtkinter.CTkButton(master=tab2_frame, text="REFRESH", command=lambda: refresh(excel_path))
button_refresh.pack(pady=10)

process_excel(excel_path)

window.mainloop()
